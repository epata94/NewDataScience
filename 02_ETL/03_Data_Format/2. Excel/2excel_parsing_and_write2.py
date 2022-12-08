# xlwt 모듈 설치
# 목적: 단일 워크시트 처리

from openpyxl import load_workbook
from openpyxl import Workbook

input_file = 'sales_2013.xlsx'
output_file = 'output_files/2output_basic.xlsx'

output_workbook = Workbook()
output_worksheet = output_workbook.create_sheet('jan_2013_output')

workbook =  load_workbook(input_file)

worksheet = workbook['january_2013']
for row_index in range(1, worksheet.max_row+1):
	for column_index in range(1,worksheet.max_column+1):
		worksheet.cell(row=3, column=2).value
		cell_value = worksheet.cell(row=row_index, column=column_index).value
		output_worksheet.cell(row_index, column_index, cell_value)

output_workbook.remove(output_workbook['Sheet'])
output_workbook.save(output_file)