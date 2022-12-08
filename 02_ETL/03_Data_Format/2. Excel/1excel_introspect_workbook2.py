# xlrd 모듈 설치
# 목적: 엑셀 기본 정보 확인
from xlrd import open_workbook
from openpyxl import load_workbook

input_file = 'sales_2013.xlsx'

workbook = load_workbook(input_file)
print('Number of worksheets:', len(workbook.worksheets))
for worksheet in workbook.worksheets:
	print("Worksheet name:", worksheet.title, "\tRows:", \
			worksheet.max_row, "\tColumns:", worksheet.max_column)
