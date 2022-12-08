# xlwt 모듈 설치
# 목적: 단일 워크시트 처리

from openpyxl import load_workbook
from openpyxl import Workbook

input_file = 'sales_2013.xlsx'

output_workbook = Workbook()
output_worksheet = output_workbook.create_sheet('jan_2013_output')

workbook =  load_workbook(input_file)

worksheet = workbook['january_2013']

# error
# print(worksheet.cell(row=0, column=1).value)

print(worksheet.cell(row=1, column=1).value)
print(worksheet.cell(row=2, column=1).value)
print(worksheet.cell(row=3, column=1).value)

print(worksheet.cell(row=1, column=2).value)
print(worksheet.cell(row=2, column=2).value)
print(worksheet.cell(row=3, column=2).value)

